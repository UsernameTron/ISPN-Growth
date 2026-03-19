"""Generate ranked Excel reports from scored partner data."""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from src.config import config

# Color definitions
GREEN_FILL = PatternFill(start_color="38A169", end_color="38A169", fill_type="solid")
AMBER_FILL = PatternFill(start_color="D69E2E", end_color="D69E2E", fill_type="solid")
RED_FILL = PatternFill(start_color="E53E3E", end_color="E53E3E", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SIGNAL_COLUMNS = [
    "sig_volume_growth",
    "sig_sla_degradation",
    "sig_service_concentration",
    "sig_bead_exposure",
    "sig_utilization_headroom",
    "sig_repeat_contacts",
    "sig_contract_proximity",
    "sig_seasonal_volatility",
]


def _tier_fill(tier: str) -> PatternFill:
    """Return the fill color for a tier value."""
    if tier == "green":
        return GREEN_FILL
    elif tier == "amber":
        return AMBER_FILL
    return RED_FILL


def _apply_header_style(ws: openpyxl.worksheet.worksheet.Worksheet, col_count: int) -> None:
    """Apply header styling to the first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_size_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Set column widths based on approximate max content length."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 3, 40)
        ws.column_dimensions[col_letter].width = adjusted_width


def _parse_top_signals(top_signals_str: str) -> list[str]:
    """Parse 'name=score, name=score, ...' into list of signal names."""
    parts = [p.strip() for p in str(top_signals_str).split(",") if "=" in p]
    return [p.split("=")[0].strip() for p in parts]


def _build_ranked_sheet(wb: openpyxl.Workbook, scored_df: pd.DataFrame) -> None:
    """Build Sheet 1: Ranked Partners."""
    ws = wb.active
    ws.title = "Ranked Partners"

    headers = [
        "Partner Name", "Composite Score", "Rank",
        "Top Signal 1", "Top Signal 2", "Top Signal 3",
        "Recommended Play", "Tier",
    ]
    ws.append(headers)
    _apply_header_style(ws, len(headers))

    for idx, (_, row) in enumerate(scored_df.iterrows(), start=1):
        top_sigs = _parse_top_signals(row.get("top_signals", ""))
        ws.append([
            row.get("partner_name", row.get("partner_id", "")),
            round(row["composite_score"], 2),
            idx,
            top_sigs[0] if len(top_sigs) > 0 else "",
            top_sigs[1] if len(top_sigs) > 1 else "",
            top_sigs[2] if len(top_sigs) > 2 else "",
            row.get("recommended_play", ""),
            row["tier"],
        ])
        # Apply tier color to the tier cell and score cell
        data_row = idx + 1  # +1 for header
        tier_cell = ws.cell(row=data_row, column=8)
        tier_cell.fill = _tier_fill(row["tier"])
        tier_cell.font = Font(color="FFFFFF", bold=True)
        score_cell = ws.cell(row=data_row, column=2)
        score_cell.fill = _tier_fill(row["tier"])
        score_cell.font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    _auto_size_columns(ws)


def _build_signal_detail_sheet(wb: openpyxl.Workbook, scored_df: pd.DataFrame) -> None:
    """Build Sheet 2: Signal Detail."""
    ws = wb.create_sheet("Signal Detail")

    # Friendly signal names for headers
    signal_display = {
        "sig_volume_growth": "Volume Growth",
        "sig_sla_degradation": "SLA Degradation",
        "sig_service_concentration": "Service Concentration",
        "sig_bead_exposure": "BEAD Exposure",
        "sig_utilization_headroom": "Utilization Headroom",
        "sig_repeat_contacts": "Repeat Contacts",
        "sig_contract_proximity": "Contract Proximity",
        "sig_seasonal_volatility": "Seasonal Volatility",
    }

    headers = ["Partner Name"]
    for sig_col in SIGNAL_COLUMNS:
        headers.append(signal_display.get(sig_col, sig_col))
    headers.extend(["Composite Score", "Tier"])

    ws.append(headers)
    _apply_header_style(ws, len(headers))

    for idx, (_, row) in enumerate(scored_df.iterrows(), start=1):
        values = [row.get("partner_name", row.get("partner_id", ""))]
        for sig_col in SIGNAL_COLUMNS:
            values.append(int(row.get(sig_col, 0)))
        values.append(round(row["composite_score"], 2))
        values.append(row["tier"])
        ws.append(values)

        # Apply tier color
        data_row = idx + 1
        tier_col = len(headers)
        score_col = tier_col - 1
        ws.cell(row=data_row, column=tier_col).fill = _tier_fill(row["tier"])
        ws.cell(row=data_row, column=tier_col).font = Font(color="FFFFFF", bold=True)
        ws.cell(row=data_row, column=score_col).fill = _tier_fill(row["tier"])
        ws.cell(row=data_row, column=score_col).font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "A2"
    _auto_size_columns(ws)


def _build_methodology_sheet(wb: openpyxl.Workbook) -> None:
    """Build Sheet 3: Methodology."""
    ws = wb.create_sheet("Methodology")

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)

    # Title
    ws.append(["GCS Engine — Methodology"])
    ws.cell(row=1, column=1).font = title_font
    ws.append([])

    # Weight table
    ws.append(["Signal Weights"])
    ws.cell(row=3, column=1).font = section_font
    ws.append(["Signal", "Weight", "Description"])
    _apply_header_style(ws, 3)
    row_num = 5
    for sw in config.signal_weights:
        ws.append([sw.name, f"{sw.weight:.0%}", sw.description])
        row_num += 1

    ws.append([])
    ws.append(["Total (active)", f"{config.total_weight:.0%}", ""])
    ws.append([])

    # Scoring criteria
    ws.append(["Scoring Criteria (0-3 per signal)"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.append(["Score", "Meaning"])
    ws.append([0, "No signal / missing data / below threshold"])
    ws.append([1, "Low signal — minor presence"])
    ws.append([2, "Moderate signal — clear presence"])
    ws.append([3, "Strong signal — high opportunity indicator"])
    ws.append([])

    # Tier definitions
    ws.append(["Tier Definitions"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.append(["Tier", "Score Range", "Meaning"])
    ws.append(["Green", "> 70", "High growth opportunity"])
    ws.append(["Amber", "40 - 70", "Moderate opportunity"])
    ws.append(["Red", "< 40", "Low opportunity"])
    ws.append([])

    # Data sources
    ws.append(["Data Sources"])
    ws.cell(row=ws.max_row, column=1).font = section_font
    ws.append(["Source", "Signals Derived"])
    ws.append(["Genesys Cloud CX", "volume_growth, sla_degradation"])
    ws.append(["HelpDesk Ticketing", "repeat_contacts"])
    ws.append(["UKG Workforce Mgmt", "utilization_headroom"])
    ws.append(["WCS Contact Summary", "seasonal_volatility"])
    ws.append(["Service Mix (CSV)", "service_concentration"])
    ws.append(["BEAD Status (CSV)", "bead_exposure"])
    ws.append([])

    # Generated date
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])

    _auto_size_columns(ws)


def generate_report(
    scored_df: pd.DataFrame,
    output_dir: str = "output",
    output_format: str = "xlsx",
    **kwargs: str,
) -> str:
    """Generate ranked report from scored partner data.

    Args:
        scored_df: DataFrame from score_all_partners with composite scores and tiers.
        output_dir: Directory to write output files.
        output_format: "xlsx", "csv", or "both".
        **kwargs: Accepts deprecated ``format`` keyword for backward compatibility.

    Returns:
        Path to the primary output file.
    """
    # Backward compatibility: accept 'format' kwarg (shadows builtin, hence renamed)
    if "format" in kwargs:
        output_format = kwargs.pop("format")
    os.makedirs(output_dir, exist_ok=True)
    date_suffix = datetime.now().strftime("%Y-%m")
    paths = []

    if output_format in ("xlsx", "both"):
        wb = openpyxl.Workbook()
        _build_ranked_sheet(wb, scored_df)
        _build_signal_detail_sheet(wb, scored_df)
        _build_methodology_sheet(wb)

        xlsx_path = os.path.join(output_dir, f"gcs_report_{date_suffix}.xlsx")
        wb.save(xlsx_path)
        paths.append(xlsx_path)

    if output_format in ("csv", "both"):
        # Export Sheet 1 equivalent as CSV
        top_sigs_parsed = scored_df["top_signals"].apply(_parse_top_signals)
        csv_df = pd.DataFrame({
            "partner_name": scored_df.get("partner_name", scored_df.get("partner_id", "")),
            "composite_score": scored_df["composite_score"].round(2),
            "rank": range(1, len(scored_df) + 1),
            "top_signal_1": top_sigs_parsed.apply(lambda x: x[0] if len(x) > 0 else ""),
            "top_signal_2": top_sigs_parsed.apply(lambda x: x[1] if len(x) > 1 else ""),
            "top_signal_3": top_sigs_parsed.apply(lambda x: x[2] if len(x) > 2 else ""),
            "recommended_play": scored_df["recommended_play"],
            "tier": scored_df["tier"],
        })
        csv_path = os.path.join(output_dir, f"gcs_report_{date_suffix}.csv")
        csv_df.to_csv(csv_path, index=False)
        paths.append(csv_path)

    return paths[0] if paths else ""
