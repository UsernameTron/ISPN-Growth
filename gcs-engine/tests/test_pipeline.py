"""End-to-end pipeline tests for the GCS Engine."""

import os
import shutil
import tempfile

import openpyxl
import pandas as pd
import pytest

from src.connectors import (
    GenesysConnector, HelpDeskConnector, UKGConnector,
    WCSConnector, ServiceMixConnector, BEADConnector,
)
from src.scoring.engine import score_all_partners
from src.output.report import generate_report
from src.output.summary import print_summary, generate_markdown_summary


@pytest.fixture
def output_dir():
    d = tempfile.mkdtemp()
    yield d
    shutil.rmtree(d)


@pytest.fixture
def full_scored_df():
    """Run the full scoring pipeline with stub connectors."""
    genesys_df = GenesysConnector().get_data()
    helpdesk_df = HelpDeskConnector().get_data()
    ukg_df = UKGConnector().get_data()
    wcs_df = WCSConnector().get_data()
    service_mix_df = ServiceMixConnector().get_data()
    bead_df = BEADConnector().get_data()

    scored_df = score_all_partners(
        genesys_df, helpdesk_df, ukg_df, wcs_df, service_mix_df, bead_df,
    )

    # Add partner_name from service_mix
    if "partner_name" in service_mix_df.columns:
        name_map = service_mix_df.set_index("partner_id")["partner_name"]
        scored_df["partner_name"] = scored_df["partner_id"].map(name_map).fillna(scored_df["partner_id"])
    else:
        scored_df["partner_name"] = scored_df["partner_id"]

    return scored_df


class TestFullPipeline:
    def test_scored_df_has_30_rows(self, full_scored_df):
        assert len(full_scored_df) == 30

    def test_all_composite_scores_in_range(self, full_scored_df):
        assert full_scored_df["composite_score"].min() >= 0.0
        assert full_scored_df["composite_score"].max() <= 100.0

    def test_tier_distribution_sensible(self, full_scored_df):
        tier_counts = full_scored_df["tier"].value_counts()
        # All tiers should be valid
        assert set(tier_counts.index).issubset({"green", "amber", "red"})
        # Should have at least some partners in each tier (with 30 partners)
        assert tier_counts.sum() == 30

    def test_excel_output_created(self, full_scored_df, output_dir):
        path = generate_report(full_scored_df, output_dir, format="xlsx")
        assert os.path.exists(path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Ranked Partners"]
        # Header + 30 data rows
        assert ws.max_row == 31

    def test_markdown_output_created(self, full_scored_df, output_dir):
        path = generate_markdown_summary(full_scored_df, output_dir)
        assert os.path.exists(path)
        with open(path) as f:
            content = f.read()
        assert "Total partners scored" in content
        assert "30" in content

    def test_partners_flag_limits_output(self, full_scored_df, output_dir):
        limited = full_scored_df.head(10)
        path = generate_report(limited, output_dir, format="xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb["Ranked Partners"]
        assert ws.max_row == 11  # header + 10

    def test_csv_output_created(self, full_scored_df, output_dir):
        path = generate_report(full_scored_df, output_dir, format="csv")
        assert os.path.exists(path)
        df = pd.read_csv(path)
        assert len(df) == 30

    def test_print_summary_runs(self, full_scored_df, capsys):
        print_summary(full_scored_df)
        captured = capsys.readouterr()
        assert "Total partners scored: 30" in captured.out
        assert "TIER DISTRIBUTION" in captured.out
