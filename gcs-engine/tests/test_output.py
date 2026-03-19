"""Tests for output report and summary generation."""

import os
import shutil
import tempfile

import openpyxl
import pandas as pd
import pytest

from src.output.report import generate_report, SIGNAL_COLUMNS
from src.output.summary import print_summary, generate_markdown_summary


def _make_scored_df(n: int = 5) -> pd.DataFrame:
    """Create a minimal scored DataFrame for testing."""
    rows = []
    for i in range(n):
        score = 90 - i * 20  # 90, 70, 50, 30, 10
        if score > 70:
            tier = "green"
        elif score >= 40:
            tier = "amber"
        else:
            tier = "red"
        row = {
            "partner_id": f"P{i + 1:03d}",
            "partner_name": f"Partner {i + 1}",
            "composite_score": float(score),
            "tier": tier,
            "top_signals": "volume_growth=3, sla_degradation=2, bead_exposure=1",
            "recommended_play": "General expansion conversation",
            "missing_signals": "",
        }
        for sig_col in SIGNAL_COLUMNS:
            row[sig_col] = max(0, 3 - i)
        rows.append(row)
    return pd.DataFrame(rows)


@pytest.fixture
def output_dir():
    """Create a temporary directory for test output."""
    d = tempfile.mkdtemp()
    yield d
    shutil.rmtree(d)


@pytest.fixture
def scored_df():
    return _make_scored_df()


class TestGenerateReport:
    def test_creates_xlsx_file(self, scored_df, output_dir):
        path = generate_report(scored_df, output_dir, format="xlsx")
        assert path.endswith(".xlsx")
        assert os.path.exists(path)

    def test_xlsx_has_three_sheets(self, scored_df, output_dir):
        path = generate_report(scored_df, output_dir, format="xlsx")
        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) == 3
        assert "Ranked Partners" in wb.sheetnames
        assert "Signal Detail" in wb.sheetnames
        assert "Methodology" in wb.sheetnames

    def test_ranked_sheet_has_correct_columns(self, scored_df, output_dir):
        path = generate_report(scored_df, output_dir, format="xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb["Ranked Partners"]
        headers = [cell.value for cell in ws[1]]
        assert "Partner Name" in headers
        assert "Composite Score" in headers
        assert "Rank" in headers
        assert "Top Signal 1" in headers
        assert "Top Signal 2" in headers
        assert "Top Signal 3" in headers
        assert "Recommended Play" in headers
        assert "Tier" in headers

    def test_signal_detail_has_all_signal_columns(self, scored_df, output_dir):
        path = generate_report(scored_df, output_dir, format="xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb["Signal Detail"]
        headers = [cell.value for cell in ws[1]]
        # Should have 8 signal columns plus Partner Name, Composite Score, Tier
        expected_signals = [
            "Volume Growth", "SLA Degradation", "Service Concentration",
            "BEAD Exposure", "Utilization Headroom", "Repeat Contacts",
            "Contract Proximity", "Seasonal Volatility",
        ]
        for sig in expected_signals:
            assert sig in headers, f"Missing signal column: {sig}"

    def test_csv_output_mode(self, scored_df, output_dir):
        path = generate_report(scored_df, output_dir, format="csv")
        assert path.endswith(".csv")
        assert os.path.exists(path)
        df = pd.read_csv(path)
        assert "partner_name" in df.columns
        assert "composite_score" in df.columns
        assert len(df) == len(scored_df)

    def test_both_output_mode(self, scored_df, output_dir):
        path = generate_report(scored_df, output_dir, format="both")
        # Returns the xlsx path first
        assert path.endswith(".xlsx")
        assert os.path.exists(path)
        # CSV should also exist
        csv_path = path.replace(".xlsx", ".csv")
        assert os.path.exists(csv_path)

    def test_ranked_sheet_row_count(self, scored_df, output_dir):
        path = generate_report(scored_df, output_dir, format="xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb["Ranked Partners"]
        # Header + data rows
        assert ws.max_row == len(scored_df) + 1


class TestMarkdownSummary:
    def test_creates_markdown_file(self, scored_df, output_dir):
        path = generate_markdown_summary(scored_df, output_dir)
        assert path.endswith(".md")
        assert os.path.exists(path)

    def test_markdown_contains_key_sections(self, scored_df, output_dir):
        path = generate_markdown_summary(scored_df, output_dir)
        with open(path) as f:
            content = f.read()
        assert "Tier Distribution" in content
        assert "Top" in content
        assert "Partner" in content


class TestPrintSummary:
    def test_does_not_crash_on_empty_df(self, capsys):
        empty_df = pd.DataFrame(columns=[
            "partner_id", "composite_score", "tier",
            "top_signals", "recommended_play",
        ])
        print_summary(empty_df)
        captured = capsys.readouterr()
        assert "No partners to summarize" in captured.out

    def test_prints_output_for_valid_df(self, scored_df, capsys):
        print_summary(scored_df)
        captured = capsys.readouterr()
        assert "TIER DISTRIBUTION" in captured.out
        assert "TOP" in captured.out
        assert "Total partners scored: 5" in captured.out
